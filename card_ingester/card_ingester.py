import xml.etree.ElementTree as eTree
import json
import traceback
import csv
from typing import List
import jaydebeapi
from jaydebeapi import Error
from openpyxl import load_workbook
from openpyxl import worksheet

class Card:
    def __init__(self):
            self.acc = None
            self.num = None
            self.pin = None
            self.cvc1 = None
            self.cvc2 = None
            self.exp = None

def parse_json_dict(json_dict: dict) -> Card:

    card = Card()
    card.acc = json_dict["accounts_id"]
    card.num = json_dict["card_num"]
    card.pin = json_dict["pin"]
    card.cvc1 = json_dict["cvc1"]
    card.cvc2 = json_dict["cvc2"]
    card.exp = json_dict["exp_date"]
    return card

def parse_file_json(path) -> List:
    try:
        f = open(path, "r")
    except:
        print("error opening file")
        return None

    json_list = json.load(f)
    return_list = []
    for json_dict in json_list:
        try:
            return_list.append(parse_json_dict(json_dict))
        except:
            print("Line failed to be parsed")
    return return_list

def parse_table_xlsx(ws: worksheet, bounds: tuple) -> List:
    ret_list = []
    for row in ws.iter_rows(min_row=bounds[0], min_col=bounds[1], max_col=bounds[1]+7):
        
        try: #test to see if we've reached the end of our data
            if row[0].value == None:
                return ret_list
        except:
            return ret_list
        try:
            card = Card()
            card.acc = row[0].value
            card.num = row[1].value
            if row[2].value == '\\N':
                card.pin = None
            else:
                card.pin = row[2].value
            card.cvc1 = row[3].value
            card.cvc2 = row[4].value
            card.exp = row[5].value.date()
            ret_list.append(card)
        except Error as e:
            print('Row failed to be parsed')
            print (e)
    return ret_list

def find_xlsx_bounds(ws: worksheet):
    row_num = 0
    for row in ws.iter_rows(min_row=1, max_row=20):
        row_num += 1 #yes this is an odd pattern, but it lets me use the iterator and report back the row number
        for i in range(0, 10):
            try:
                if row[i].value == "id": #header and primary key
                    return (row_num+1, i+2)# row_num is already 1 indexed, but we want to add one because we hit id. For column we add one for 0-index and one to get from id to cards_id
                elif row[i].value == "cards_id":
                    return (row_num+1, i+1)#adjust row_num for headers and i for 0-index
                elif row[i].value != '' and row[i].value != None: #if we hit data
                    try:
                        if row[i+8].value == '' or row[i+8].value == None: #if the 9th col is empty we assume it doesn't have primary key
                            return(row_num, i+1) #adjust i for 0-index
                        else:
                            return(row_num, i+2) #adjust i for 0-index and primary key
                    except: #means there's probably not anything there
                            return(row_num, i+1) #adjust i for 0-index
            except:
                pass #we don't care about an out of range here
    return None

def parse_file_xlsx(path: str) -> List:
    wb = load_workbook(filename=path, read_only=True)
    ws = None
    for sheet in wb: #tries to find a sheet named cards
        if sheet.title.lower() == "cards":
            ws = sheet
    if ws == None:
        ws = wb.active #gets the first sheet
    bounds = find_xlsx_bounds(ws)
    return parse_table_xlsx(ws, bounds)

def date_to_string(date): #differs from str(date) in that it accepts none and cuts off the time component
    if date:
        return str(date)        
    return None

def write_cards(cards: List, conn: jaydebeapi.Connection) -> None:
    query = 'INSERT INTO cards(accounts_id, card_num, pin, cvc1, cvc2, exp_date) VALUES (?,?,?,?,?,?)'
    cur = conn.cursor()
    for card in cards:
        vals = (card.acc, card.num, card.pin, card.cvc1, card.cvc2, date_to_string(card.exp))
        try:
            cur.execute(query, vals)
        except jaydebeapi.Error as e:
            print("could not write card with values " + str(vals))
            print(e)

def parse_file_csv(path: str) -> List:
    with open(path, newline="") as file:
        reader = csv.reader(file, delimiter=',')
        row_count = 0
        ret_list = [] #return list
        for row in reader:
            # Functionality for ingesting cards
            try:
                card = Card()
                card.acc = row[0]
                card.num = row[1]
                if row[2] == '\\N':
                    card.pin = None
                else:
                    card.pin = row[2]
                card.cvc1 = row[3]
                card.cvc2 = row[4]
                card.exp = row[5]
                ret_list.append(card)
            except:
                print("Could not add card on line " + str(row_count) + ": " + str(row))
                continue
    return ret_list

def parse_file_xml(path: str) -> List:
    ret_list = []
    try:
        tree = eTree.parse(path)
        root = tree.getroot()
        for child in root:
            try:
                card = Card()
                card.acc = child.find('accounts_id').text
                card.num = child.find('card_num').text
                card.pin = child.find('pin').text
                card.cvc1 = child.find('cvc1').text
                card.cvc2 = child.find('cvc2').text
                card.exp = child.find('exp_date').text
                ret_list.append(card)
            except:
                print("Could not add user")
                print("Skipping line...\n")
    except eTree.ParseError:
        traceback.print_exc()
    return ret_list

def connect_mysql(path):
    con_try = None
    try:
        f = open(path, 'r')
        key = json.load(f)        
        con_try = jaydebeapi.connect(key["driver"], key["location"], key["login"], key["jar"] )
        con_try.jconn.setAutoCommit(False)
    except Error:
        print("There was a problem connecting to the database, please make sure the database information is correct!")
    return con_try

def read_file(path, conn): #ENTRY POINT
    try:
        ending = path.split('.')[-1].lower()
    except:
        print("Error finding file ending")
        return
    if ending == "json":
        card = parse_file_json(path)
    elif ending == "xlsx":
        card = parse_file_xlsx(path)
    elif ending == "csv":
        card = parse_file_csv(path)
    else:
        print ("Invalid file format")
        return
    write_cards(card, conn)
