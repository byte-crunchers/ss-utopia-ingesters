import csv
import json
import os
import traceback
from typing import List
import xml.etree.ElementTree
import xml.etree.ElementTree as eTree
import jaydebeapi
import mysql
from mysql.connector import Error
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
num_of_fields = 1


class Branch:
    def __init__(self, street_address, city, state, zip_code):
        self.street_address = street_address
        self.city = city
        self.state = state
        self.zip_code = zip_code

    def print_branch(self):
        print(self.street_address, self.city, self.state, self.zip_code)


def populate_branches(branch_data, ing_conn):
    curs = ing_conn.cursor()
    query = "INSERT INTO branches(street_address, city, state, zip) VALUES(?, ?, ?, ?) "
    for branch in branch_data:
        try:
            # Checking if any of the incoming values are null or duplicate, if they are skip addition
            if branch.street_address and branch.street_address.strip() and branch.city and branch.city.strip()\
                    and branch.state and branch.state.strip() and branch.zip_code and branch.zip_code.strip():
                vals = (branch.street_address, branch.city, branch.state, branch.zip_code)
                curs.execute(query, vals)
            else:
                raise Exception
            # Check for Duplicates and Nulls
        except (mysql.connector.errors.IntegrityError, jaydebeapi.DatabaseError, Exception):
            print("Duplicate Location or Illegal Null: ")
            branch.print_branch()
            print("Skipping addition..\n")


# Parse csv into users
def parse_file_csv(file):
    branch_list = []
    with open(file, newline="") as branch_file:
        reader = csv.reader(branch_file, delimiter=',')
        row_count = 0
        for row in reader:
            # Functionality for ingesting branches
            try:
                branch_list.append(
                    Branch(str(row[1]), str(row[2]), str(row[3]), str(row[4]))
                )
                row_count += 1
            except (ValueError, IndexError, Exception):
                print("Could not add branch on line " + str(row_count) + ": " + str(row))
                print("Skipping line...\n")
                continue
    return branch_list


def parse_file_xml(path):
    xml_branch_list = []
    try:
        tree = eTree.parse(path)
        root = tree.getroot()
        for child in root:
            try:
                xml_branch_list.append(
                    Branch(child.find('id').text, child.find('location').text))
            except ValueError:
                print("Could not add branch:" + child.find('id').text)
                print("Skipping line...\n")
    except xml.etree.ElementTree.ParseError:
        traceback.print_exc()
    return xml_branch_list


def parse_json_dict(json_dict: dict) -> Branch:
    try:
        json_branch = Branch(json_dict["id"], json_dict["location"])
        return json_branch
    except (ValueError, IndexError):
        print("Could not add branch: " + json_dict["id"], json_dict["location"])
        print("Skipping line...\n")


def parse_file_json(path) -> List:
    f = None
    try:
        f = open(path, "r")
    except IOError:
        print("error opening file")
    json_list = json.load(f)
    return_list = []
    for json_dict in json_list:
        return_list.append(parse_json_dict(json_dict))
    return return_list


def parse_table_xlsx(ws: worksheet, bounds: tuple) -> List:
    ret_list = []
    for row in ws.iter_rows(min_row=bounds[0], min_col=bounds[1], max_col=bounds[1] + 7):
        try:  # test to see if we've reached the end of our data
            if row[0].value is None:
                return ret_list
        except IndexError:
            return ret_list
        branch = Branch(0, row[0].value)
        ret_list.append(branch)
    return ret_list


def find_xlsx_bounds(ws: worksheet):
    row_num = 0
    for row in ws.iter_rows(min_row=1, max_row=20):
        row_num += 1  # yes this is an odd pattern, but it lets me use the iterator and report back the row number
        for i in range(0, 10):
            try:
                if row[i].value == "id":  # header and primary key
                    return (row_num + 1,
                            i + 2)  # row_num is already 1 indexed, but we want to add one because we hit id. For
                elif row[i].value == "location":
                    return row_num + 1, i + 1  # adjust row_num for headers and i for 0-index
                elif row[i].value != '' and row[i].value is not None:  # if we hit data
                    try:
                        # if the 9th col is empty we assume
                        if row[i + num_of_fields].value == '' or row[i + num_of_fields].value is None:
                            # it doesn't have primary key
                            return row_num, i + 1  # adjust i for 0-index
                        else:
                            return row_num, i + 2  # adjust i for 0-index and primary key
                    except:  # means there's probably not anything there
                        return row_num, i + 1  # adjust i for 0-index
            except:
                pass  # we don't care about an out of range here
    return None


def parse_file_xlsx(path: str) -> List:
    wb = load_workbook(filename=path, read_only=True)
    ws = None
    for sheet in wb:  # tries to find a sheet named accounts
        if sheet.title.lower() == "accounts":
            ws = sheet
    if ws is None:
        ws = wb.active  # gets the first sheet
    bounds = find_xlsx_bounds(ws)
    return parse_table_xlsx(ws, bounds)


# ENTRY POINT
def read_file(path, conn):
    try:
        ending = path.split('.')[-1].lower()
    except:
        print("Error finding file ending")
        return
    if ending == "json":
        acc = parse_file_json(path)
    elif ending == "xlsx":
        acc = parse_file_xlsx(path)
    elif ending == "csv":
        acc = parse_file_csv(path)
    elif ending == "xml":
        acc = parse_file_xml(path)
    else:
        print("Invalid file format")
        return
    populate_branches(acc, conn)


def clear_table(table, clear_conn):
    queries = []
    h2_query = "DELETE FROM {};".format(table)
    queries.append(h2_query)
    try:
        clear_curs = clear_conn.cursor()
        for q in queries:
            clear_curs.execute(q)
    except Exception:
        print("There was a problem clearing the table!")


# This returns the count of all rows in the table
def count_rows(table, count_conn):
    count_curs = count_conn.cursor()
    count_query = "select count(*) from {}".format(table)
    row_count = None
    try:
        count_curs.execute(count_query)
        row_count = count_curs.fetchall()[0][0]
    except Error:
        traceback.print_exc(
            print("There was a problem counting the rows")
        )
    return row_count


def execute_scripts_from_file(filename, conn):
    # Open and read the file as a single buffer
    sql_file = None
    try:
        fd = open(filename, 'r')
        sql_file = fd.read()
        fd.close()
    except IOError:
        print("Error opening sql file...\n")
        return None
    # all sql commands (split on ';')
    sql_commands = sql_file.split(';')
    # Execute every command from the input file
    curs = conn.cursor()
    for command in sql_commands:
        # This will skip and report errors
        # For example, if the tables do not yet exist, this will skip over
        # the DROP TABLE commands
        try:
            curs.execute(command)
        except (jaydebeapi.OperationalError, jaydebeapi.DatabaseError, Exception):
            if command.isspace():
                print("Skipping blank command in sql...\n")
            else:
                print("Could not execute command: ", command, "skipping command...\n")


