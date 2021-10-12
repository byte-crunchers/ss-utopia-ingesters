import csv
import json
import traceback
from typing import List
import xml.etree.ElementTree
import xml.etree.ElementTree as eTree
import jaydebeapi
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet


class LoanPayment:
    def __init__(self, loan_id, account_id, amount, time_stamp, status):
        self.loan_id = loan_id
        self.account_id = account_id
        self.amount = amount
        self.time_stamp = time_stamp
        self.status = status

    def print_user(self):
        try:
            print(
                str(self.loan_id) + ", " + str(self.account_id) + ", " + str(self.time_stamp) + ", " + str(self.status))
        except TypeError:
            print("Illegal Null Found")


def populate_loan_payments(lp_data, ing_conn):
    row_count = 0
    curs = ing_conn.cursor()
    dd_count = 0
    curs = ing_conn.cursor()
    query = "INSERT INTO loan_payments(loan_id, account_id, amount, time_stamp, status) VALUES(?, ?, ?, ?, ?) "
    for lp in lp_data:
        row_count += 1
        try:
            if str(lp.loan_id).strip() and str(lp.account_id).strip() and str(lp.amount).strip() and \
                    str(lp.status).strip():
                vals = (lp.loan_id, lp.account_id, lp.amount, lp.time_stamp, lp.status)
                curs.execute(query, vals)
            else:
                raise Exception
            # Check for Duplicates and Nulls
        except jaydebeapi.DatabaseError:
            print("Illegal Null on row " + str(row_count) + ":")
            lp.print_user()
            print("Skipping addition..\n")
        except Exception:
            traceback.print_exc()


# Parse csv into users
def parse_file_csv(file):
    lp_list = []
    with open(file, newline="") as user_file:
        reader = csv.reader(user_file, delimiter=',')
        row_count = 0
        for row in reader:
            # Functionality for ingesting
            try:
                lp_list.append(
                    LoanPayment(str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]))
                )
                row_count += 1
            except (ValueError, IndexError):
                print("Could not add on line " + str(row_count) + ": " + str(row))
                print("Skipping line...\n")
                continue
    return lp_list


def parse_json_dict(json_dict: dict) -> LoanPayment:
    json_lp = LoanPayment(json_dict["loan_id"], json_dict["account_id"], json_dict["amount"],
                          json_dict["time_stamp"], json_dict["status"])
    return json_lp


def parse_file_json(path) -> List:
    f = None
    try:
        f = open(path, "r")
    except IOError:
        print("error opening file")
    json_list = json.load(f)
    return_list = []
    for json_dict in json_list:
        return_list.append(parse_json_dict(json_dict))
    return return_list


def parse_file_xml(path):
    xml_user_list = []
    try:
        tree = eTree.parse(path)
        root = tree.getroot()
        for child in root:
            try:
                xml_user_list.append(
                    LoanPayment(child.find('loan_id').text, child.find('account_id').text,
                                child.find('amount').text, child.find('time_stamp').text, child.find('status').text)
                )

            except ValueError:
                print("Could not add user:" + child.find('id').text)
                print("Skipping line...\n")
    except xml.etree.ElementTree.ParseError:
        traceback.print_exc()
    return xml_user_list


def parse_table_xlsx(ws: worksheet, bounds: tuple) -> List:
    ret_list = []
    for row in ws.iter_rows(min_row=bounds[0], min_col=bounds[1], max_col=bounds[1] + 5):
        try:  # test to see if we've reached the end of our data
            if row[0].value is None:
                return ret_list
        except IndexError:
            return ret_list
        loan_payment = LoanPayment(row[0].value, row[1].value, row[2].value, str(row[3].value), row[4].value)
        ret_list.append(loan_payment)
    return ret_list


def find_xlsx_bounds(ws: worksheet):
    # Actually number of fields - 1
    num_of_fields = 4
    second_field = "loan_id"
    row_num = 0
    for row in ws.iter_rows(min_row=1, max_row=20):
        row_num += 1  # yes this is an odd pattern, but it lets me use the iterator and report back the row number
        for i in range(0, 10):
            try:
                if row[i].value == "id":  # header and primary key
                    return (row_num + 1,
                            i + 2)  # row_num is already 1 indexed, but we want to add one because we hit id. For
                    # column we add one for 0-index and one to get from id to accounts_id
                elif row[i].value == second_field:
                    return row_num + 1, i + 1  # adjust row_num for headers and i for 0-index
                elif row[i].value != '' and row[i].value is not None:  # if we hit data
                    try:
                        # if the 9th col is empty we assume
                        if row[i + num_of_fields].value == '' or row[i + num_of_fields].value is None:
                            # it doesn't have primary key
                            return row_num, i + 1  # adjust i for 0-index
                        else:
                            return row_num, i + 2  # adjust i for 0-index and primary key
                    except:  # means there's probably not anything there
                        return row_num, i + 1  # adjust i for 0-index
            except:
                pass  # we don't care about an out of range here
    return None


def parse_file_xlsx(path: str) -> List:
    wb = load_workbook(filename=path, read_only=True)
    ws = None
    for sheet in wb:  # tries to find a sheet named accounts
        if sheet.title.lower() == "users":
            ws = sheet
    if ws is None:
        ws = wb.active  # gets the first sheet
    bounds = find_xlsx_bounds(ws)
    return parse_table_xlsx(ws, bounds)


def read_file(path, conn):
    try:
        ending = path.split('.')[-1].lower()
    except:
        print("Error finding file ending")
        return
    if ending == "json":
        acc = parse_file_json(path)
    elif ending == "xlsx":
        acc = parse_file_xlsx(path)
    elif ending == "csv":
        acc = parse_file_csv(path)
    elif ending == "xml":
        acc = parse_file_xml(path)
    else:
        print("Invalid file format")
        return
    populate_loan_payments(acc, conn)


def execute_scripts_from_file(filename, conn):
    # Open and read the file as a single buffer
    sql_file = None
    try:
        fd = open(filename, 'r')
        sql_file = fd.read()
        fd.close()
    except IOError:
        print("Error opening sql file...\n")
        return None
    # all sql commands (split on ';')
    sql_commands = sql_file.split(';')
    # Execute every command from the input file
    curs = conn.cursor()
    for command in sql_commands:
        # This will skip and report errors
        # For example, if the tables do not yet exist, this will skip over
        # the DROP TABLE commands
        try:
            curs.execute(command)
        except (jaydebeapi.OperationalError, jaydebeapi.DatabaseError, Exception):
            if command.isspace():
                print("Skipping blank command in sql...\n")
            else:
                print("Could not execute command: ", command, "skipping command...\n")
