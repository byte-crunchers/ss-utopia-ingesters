import csv
import json
import traceback
import xml.etree.ElementTree
import xml.etree.ElementTree as eTree
from typing import List

import jaydebeapi
from openpyxl import load_workbook
from openpyxl import worksheet

class Loan:
    def __init__(self, users_id, balance, interest_rate, due_date, payment_due, loan_type, monthly_payment, active,
                 approved, confirmed):
        self.users_id = users_id
        self.balance = balance
        self.interest_rate = interest_rate
        self.due_date = due_date
        self.payment_due = payment_due
        self.loan_type = loan_type
        self.monthly_payment = monthly_payment
        self.active = active
        self.approved = approved
        self.confirmed = confirmed

    def print_loan(self):
        try:
            print(str(self.users_id) + ", " + str(self.balance) + ", " + str(self.interest_rate) + ", " + str(
                self.due_date) + ", "
                  + str(self.payment_due) + ", " + str(self.loan_type) + ", " + str(self.monthly_payment) + ", " + str(
                self.active) + ", "
                  + str(self.approved) + ", " + str(self.confirmed))
        except TypeError:
            print("Illegal Null Found")


def populate_loans(loan_data, conn):
    query = 'INSERT INTO loans(users_id, balance, interest_rate, due_date, payment_due, loan_type, monthly_payment, \
        active, approved, confirmed) VALUES (?,?,?,?,?,?,?,?,?,?)'
    cur = conn.cursor()
    for loan in loan_data:
        try:
            vals = (loan.users_id, loan.balance, loan.interest_rate, str(loan.due_date), loan.payment_due,
                    loan.loan_type, loan.monthly_payment, loan.active, loan.approved, loan.confirmed)
            cur.execute(query, vals)
        # Check for Duplicates and Nulls
        except jaydebeapi.DatabaseError:
            print("users_id or loan_type Dependecy missing or Illegal Null: ")
            loan.print_loan()
            print("Skipping addition..\n")
        except Exception():
            traceback.print_exc()


# Parse csv into users
def parse_file_csv(file):
    loan_list = []
    with open(file, newline="") as loan_file:
        reader = csv.reader(loan_file, delimiter=',')
        row_count = 0
        for row in reader:
            row_count += 1
            # Functionality for ingesting users
            try:
                loan_list.append(
                    Loan(int(row[1]), float(row[2]), float(row[3]), str(row[4]), float(row[5]), str(row[6]),
                         float(row[7]), bool(row[8]), bool(row[9]), bool(row[10]))
                )
            except (ValueError, IndexError):
                # traceback.print_exc()
                print("Could not add loan on line " + str(row_count) + ": " + str(row))
                print("Skipping line...\n")
                continue
    return loan_list


def parse_json_dict(json_dict: dict) -> Loan:
    json_loan = Loan(json_dict["users_id"], json_dict["balance"], json_dict["interest_rate"],
                     json_dict["due_date"], json_dict["payment_due"], json_dict["loan_type"],
                     json_dict["monthly_payment"], json_dict["active"], json_dict["approved"], json_dict["confirmed"])
    return json_loan


def parse_file_json(path):
    f = None
    try:
        f = open(path, "r")
    except IOError:
        print("error opening file")
    json_list = json.load(f)
    return_list = []
    for json_dict in json_list:
        return_list.append(parse_json_dict(json_dict))
    return return_list


def parse_file_xlsx(path):
    wb = load_workbook(filename=path, read_only=True)
    ws = None
    for sheet in wb:  # tries to find a sheet named accounts
        if sheet.title.lower() == "loans":
            ws = sheet
    if ws is None:
        ws = wb.active  # gets the first sheet
    bounds = find_xlsx_bounds(ws)
    return parse_table_xlsx(ws, bounds)


def find_xlsx_bounds(ws: worksheet):
    num_of_fields = 9
    second_field = "users_id"
    row_num = 0
    for row in ws.iter_rows(min_row=1, max_row=20):
        row_num += 1  # yes this is an odd pattern, but it lets me use the iterator and report back the row number
        for i in range(0, 10):
            try:
                if row[i].value == "id":  # header and primary key
                    return (row_num + 1,
                            i + 2)  # row_num is already 1 indexed, but we want to add one because we hit id. For
                    # column we add one for 0-index and one to get from id to accounts_id
                elif row[i].value == second_field:
                    return row_num + 1, i + 1  # adjust row_num for headers and i for 0-index
                elif row[i].value != '' and row[i].value is not None:  # if we hit data
                    try:
                        # if the 9th col is empty we assume
                        if row[i + num_of_fields].value == '' or row[i + num_of_fields].value is None:
                            # it doesn't have primary key
                            return row_num, i + 1  # adjust i for 0-index
                        else:
                            return row_num, i + 2  # adjust i for 0-index and primary key
                    except:  # means there's probably not anything there
                        return row_num, i + 1  # adjust i for 0-index
            except:
                pass  # we don't care about an out of range here
    return None


def parse_table_xlsx(ws: worksheet, bounds: tuple) -> List:
    ret_list = []
    for row in ws.iter_rows(min_row=bounds[0], min_col=bounds[1], max_col=bounds[1] + 9):
        try:  # test to see if we've reached the end of our data
            if row[0].value is None:
                return ret_list
        except IndexError:
            return ret_list
        date = row[3].value
        if date:
            date = date.date()
        user = Loan(row[0].value, row[1].value, row[2].value, str(date), row[4].value, row[5].value, row[6].value,
                    row[7].value, row[8].value, row[9].value)
        ret_list.append(user)
    return ret_list


def parse_file_xml(path):
    xml_user_list = []
    try:
        tree = eTree.parse(path)
        root = tree.getroot()
        for child in root:
            try:
                xml_user_list.append(
                    Loan(child.find('users_id').text, child.find('balance').text,
                         child.find('interest_rate').text, child.find('due_date').text, child.find('payment_due').text,
                         child.find('loan_type').text, child.find('monthly_payment').text, child.find('active').text,
                         child.find('approved').text, child.find('confirmed').text)
                )
            except ValueError:
                print("Could not add user:" + child.find('id').text)
                print("Skipping line...\n")
    except xml.etree.ElementTree.ParseError:
        traceback.print_exc()
    return xml_user_list


def read_file(path, conn):
    try:
        ending = path.split('.')[-1].lower()
    except:
        print("Error finding file ending")
        return
    if ending == "json":
        acc = parse_file_json(path)
    elif ending == "xlsx":
        acc = parse_file_xlsx(path)
    elif ending == "csv":
        acc = parse_file_csv(path)
    elif ending == "xml":
        acc = parse_file_xml(path)
    else:
        print("Invalid file format")
        return
    populate_loans(acc, conn)


def execute_scripts_from_file(filename, conn):
    # Open and read the file as a single buffer
    sql_file = None
    try:
        fd = open(filename, 'r')
        sql_file = fd.read()
        fd.close()
    except IOError:
        print("Error opening sql file...\n")
        return None
    # all sql commands (split on ';')
    sql_commands = sql_file.split(';')
    # Execute every command from the input file
    curs = conn.cursor()
    for command in sql_commands:
        # This will skip and report errors
        # For example, if the tables do not yet exist, this will skip over
        # the DROP TABLE commands
        try:
            curs.execute(command)
        except (jaydebeapi.OperationalError, jaydebeapi.DatabaseError, Exception):
            if command.isspace():
                print("Skipping blank command in sql...\n")
            else:
                print("Could not execute command: ", command, "skipping command...\n")
