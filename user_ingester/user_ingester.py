import csv
import json
import traceback
import xml.etree.ElementTree
import xml.etree.ElementTree as eTree
from typing import List

import jaydebeapi
from openpyxl import load_workbook
from openpyxl import worksheet


class User:
    def __init__(self, user, email, password, f_name, l_name, is_admin, ssn, active, confirmed, phone,
                 dob, street_address, city, state, zip_code):
        self.user = user
        self.email = email
        self.password = password
        self.f_name = f_name
        self.l_name = l_name
        self.is_admin = is_admin
        self.ssn = ssn
        self.active = active
        self.confirmed = confirmed
        self.phone = phone
        self.dob = dob
        self.street_address = street_address
        self.city = city
        self.state = state
        self.zip_code = zip_code

    def print_user(self):
        try:
            print(
                str(self.user) + ", " + str(self.email) + ", " + str(self.password) + ", " + self.f_name + ", "
                + self.l_name + ", " + str(self.is_admin) + ", " + str(self.ssn) + ", " + str(self.active) + ", "
                + str(self.confirmed) + ", " + str(self.phone) + ", " + str(self.dob) + ", " + str(self.street_address)
                + ", " + str(self.city) + ", " + str(self.state) + ", " + str(self.zip_code))
        except TypeError:
            print("Illegal Null Found")


def populate_users(user_data, ing_conn):
    row_count = 0
    curs = ing_conn.cursor()
    dd_count = 0
    curs = ing_conn.cursor()
    query = "INSERT INTO users(username, email, password, first_name, last_name, is_admin, ssn, active, " \
            "confirmed, phone, dob, street_address, city, state, zip) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, " \
            "?, ?, ?) "
    for usr in user_data:
        row_count += 1
        try:
            if usr.user.strip() and usr.email.strip() and usr.password.strip() and usr.f_name.strip:
                vals = (usr.user, usr.email, usr.password, usr.f_name, usr.l_name, usr.is_admin, usr.ssn,
                        usr.active, usr.confirmed, usr.phone, usr.dob, usr.street_address, usr.city,
                        usr.state, usr.zip_code)
                curs.execute(query, vals)
            else:
                raise Exception
            # Check for Duplicates and Nulls
        except (jaydebeapi.DatabaseError, Exception):
            print("Duplicate User or Illegal Null on row " + str(row_count) + ":")
            usr.print_user()
            print("Skipping addition..\n")


# Parse csv into users
def parse_file_csv(file):
    user_list = []
    with open(file, newline="") as user_file:
        reader = csv.reader(user_file, delimiter=',')
        row_count = 0
        for row in reader:
            # Functionality for ingesting users
            try:
                user_list.append(
                    User(str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]), bool(row[6]),
                         str(int(row[7])), bool(row[8]), bool(row[9]), str(int(row[10])), str(row[11]), str(row[12]),
                         str(row[13]), str(row[14]), str(row[15]))
                )
                row_count += 1
            except (ValueError, IndexError):
                print("Could not add user on line " + str(row_count) + ": " + str(row))
                print("Skipping line...\n")
                continue
    return user_list


def parse_json_dict(json_dict: dict) -> User:
    json_user = User(json_dict["username"], json_dict["email"], json_dict["password"],
                     json_dict["first_name"], json_dict["last_name"], json_dict["is_admin"], json_dict["ssn"],
                     json_dict["active"], json_dict["confirmed"], json_dict["phone"], json_dict["dob"],
                     json_dict["street_address"], json_dict["city"], json_dict["state"], json_dict["zip"])
    return json_user


def parse_file_json(path) -> List:
    f = None
    try:
        f = open(path, "r")
    except IOError:
        print("error opening file")
    json_list = json.load(f)
    return_list = []
    for json_dict in json_list:
        return_list.append(parse_json_dict(json_dict))
    return return_list


def parse_file_xml(path):
    xml_user_list = []
    try:
        tree = eTree.parse(path)
        root = tree.getroot()
        for child in root:
            try:
                xml_user_list.append(
                    User(child.find('username').text, child.find('email').text,
                         child.find('password').text, child.find('first_name').text, child.find('last_name').text,
                         child.find('is_admin').text, child.find('ssn').text, child.find('active').text,
                         child.find('confirmed').text, child.find('phone').text, child.find('dob').text,
                         child.find('street_address').text, child.find('city').text, child.find('state').text,
                         child.find('zip').text)
                )

            except ValueError:
                print("Could not add user:" + child.find('id').text)
                print("Skipping line...\n")
    except xml.etree.ElementTree.ParseError:
        traceback.print_exc()
    return xml_user_list


def parse_table_xlsx(ws: worksheet, bounds: tuple) -> List:
    ret_list = []
    for row in ws.iter_rows(min_row=bounds[0], min_col=bounds[1], max_col=bounds[1] + 14):
        try:  # test to see if we've reached the end of our data
            if row[0].value is None:
                return ret_list
        except IndexError:
            return ret_list
        date = row[10].value
        if date:
            date = date.date()
        user = User(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value,
                    row[7].value, row[8].value, row[9].value, str(date), row[11].value, row[12].value,
                    row[13].value, row[14].value)
        ret_list.append(user)
    return ret_list


def find_xlsx_bounds(ws: worksheet):
    num_of_fields = 13
    second_field = "username"
    row_num = 0
    for row in ws.iter_rows(min_row=1, max_row=20):
        row_num += 1  # yes this is an odd pattern, but it lets me use the iterator and report back the row number
        for i in range(0, 10):
            try:
                if row[i].value == "id":  # header and primary key
                    return (row_num + 1,
                            i + 2)  # row_num is already 1 indexed, but we want to add one because we hit id. For
                    # column we add one for 0-index and one to get from id to accounts_id
                elif row[i].value == second_field:
                    return row_num + 1, i + 1  # adjust row_num for headers and i for 0-index
                elif row[i].value != '' and row[i].value is not None:  # if we hit data
                    try:
                        # if the 9th col is empty we assume
                        if row[i + num_of_fields].value == '' or row[i + num_of_fields].value is None:
                            # it doesn't have primary key
                            return row_num, i + 1  # adjust i for 0-index
                        else:
                            return row_num, i + 2  # adjust i for 0-index and primary key
                    except:  # means there's probably not anything there
                        return row_num, i + 1  # adjust i for 0-index
            except:
                pass  # we don't care about an out of range here
    return None


def parse_file_xlsx(path: str) -> List:
    wb = load_workbook(filename=path, read_only=True)
    ws = None
    for sheet in wb:  # tries to find a sheet named accounts
        if sheet.title.lower() == "users":
            ws = sheet
    if ws is None:
        ws = wb.active  # gets the first sheet
    bounds = find_xlsx_bounds(ws)
    return parse_table_xlsx(ws, bounds)


# ENTRY POINT
def read_file(path, conn):
    try:
        ending = path.split('.')[-1].lower()
    except:
        print("Error finding file ending")
        return
    if ending == "json":
        acc = parse_file_json(path)
    elif ending == "xlsx":
        acc = parse_file_xlsx(path)
    elif ending == "csv":
        acc = parse_file_csv(path)
    elif ending == "xml":
        acc = parse_file_xml(path)
    else:
        print("Invalid file format")
        return
    populate_users(acc, conn)


def execute_scripts_from_file(filename, conn):
    # Open and read the file as a single buffer
    sql_file = None
    try:
        fd = open(filename, 'r')
        sql_file = fd.read()
        fd.close()
    except IOError:
        print("Error opening sql file...\n")
        return None
    # all sql commands (split on ';')
    sql_commands = sql_file.split(';')
    # Execute every command from the input file
    curs = conn.cursor()
    for command in sql_commands:
        # This will skip and report errors
        # For example, if the tables do not yet exist, this will skip over
        # the DROP TABLE commands
        try:
            curs.execute(command)
        except (jaydebeapi.OperationalError, jaydebeapi.DatabaseError, Exception):
            if command.isspace():
                print("Skipping blank command in sql...\n")
            else:
                print("Could not execute command: ", command, "skipping command...\n")
