import shutil
import os
import time
import datetime
import json
import csv
from typing import List
import jaydebeapi
from jaydebeapi import Error
from openpyxl import load_workbook
from openpyxl import worksheet


class Account:
    def __init__(self):
            self.user = None
            self.account_type = None
            self.balance = None
            self.payment_due = None
            self.due_date = None
            self.limit = None
            self.interest = None
            self.active = None

def parse_json_dict(json_dict: dict) -> Account:
    account = Account()
    account.user = json_dict["users_id"]
    account.account_type = json_dict["account_type"]
    account.balance = json_dict["balance"]
    account.payment_due = json_dict["payment_due"]
    account.due_date = json_dict["due_date"]
    account.limit = json_dict["limit"]
    account.interest = json_dict["debt_interest"]
    account.active = json_dict["active"]
    return account

def parse_file_json(path) -> List:
    try:
        f = open(path, "r")
    except:
        print("error opening file")
        return None

    json_list = json.load(f)
    return_list = []
    for json_dict in json_list:
        return_list.append(parse_json_dict(json_dict))
    return return_list

def parse_table_xlsx(ws: worksheet, bounds: tuple) -> List:
    ret_list = []
    for row in ws.iter_rows(min_row=bounds[0], min_col=bounds[1], max_col=bounds[1]+7):
        
        try: #test to see if we've reached the end of our data
            if row[0].value == None:
                return ret_list
        except:
            return ret_list
        account = Account()
        account.user = row[0].value
        account.account_type = row[1].value
        account.balance = row[2].value
        account.payment_due = row[3].value
        if row[4].value == '\\N':
            account.due_date = None
        else:
            account.due_date = row[4].value
        if row[5].value == '\\N':
            account.limit = None
        else:
            account.limit = row[5].value
        account.interest = row[6].value
        account.active = row[7].value
        ret_list.append(account)
    return ret_list

def find_xlsx_bounds(ws: worksheet):
    row_num = 0
    for row in ws.iter_rows(min_row=1, max_row=20):
        row_num += 1 #yes this is an odd pattern, but it lets me use the iterator and report back the row number
        for i in range(0, 10):
            try:
                if row[i].value == "id": #header and primary key
                    return (row_num+1, i+2)# row_num is already 1 indexed, but we want to add one because we hit id. For column we add one for 0-index and one to get from id to accounts_id
                elif row[i].value == "accounts_id":
                    return (row_num+1, i+1)#adjust row_num for headers and i for 0-index
                elif row[i].value != '' and row[i].value != None: #if we hit data
                    try:
                        if row[i+8].value == '' or row[i+8].value == None: #if the 9th col is empty we assume it doesn't have primary key
                            return(row_num, i+1) #adjust i for 0-index
                        else:
                            return(row_num, i+2) #adjust i for 0-index and primary key
                    except: #means there's probably not anything there
                            return(row_num, i+1) #adjust i for 0-index
            except:
                pass #we don't care about an out of range here
    return None

def parse_file_xlsx(path: str) -> List:
    wb = load_workbook(filename=path, read_only=True)
    ws = None
    for sheet in wb: #tries to find a sheet named accounts
        if sheet.title.lower() == "accounts":
            ws = sheet
    if ws == None:
        ws = wb.active #gets the first sheet
    bounds = find_xlsx_bounds(ws)
    return parse_table_xlsx(ws, bounds)

def date_to_string(date): #differs from str(date) in that it accepts none
    if date:
        return str(date)
    return None

def write_accounts(accounts: List, conn: jaydebeapi.Connection) -> None:
    query = 'INSERT INTO accounts(users_id, account_type, balance, payment_due, due_date, credit_limit, debt_interest, active) VALUES (?,?,?,?,?,?,?,?)'
    cur = conn.cursor()
    for account in accounts:
        vals = (account.user, account.account_type, account.balance, account.payment_due, date_to_string(account.due_date), account.limit, account.interest, account.active)
        try:
            cur.execute(query, vals)
        except:
            print("could not write account with values " + str(vals))

def parse_file_csv(path: str) -> List:
    with open(path, newline="") as acc_file:
        reader = csv.reader(acc_file, delimiter=',')
        row_count = 0
        ret_list = [] #return list
        for row in reader:
            # Functionality for ingesting accouts
            try:
                account = Account()
                account.user = row[0]
                account.account_type = row[1]
                account.balance = row[2]
                account.payment_due = row[3]
                if row[4] == '\\N':
                    account.due_date = None
                else:
                    account.due_date = row[4]
                if row[5] == '\\N':
                    account.limit = None
                else:
                    account.limit = row[5]
                account.interest = row[6]
                account.active = row[7]
                ret_list.append(account)
            except:
                print("Could not add account on line " + str(row_count) + ": " + str(row))
                continue
    return ret_list

def connect_mysql(path):
    con_try = None
    try:
        f = open(path, 'r')
        key = json.load(f)        
        con_try = jaydebeapi.connect(key["driver"], key["location"], key["login"], key["jar"] )
        con_try.jconn.setAutoCommit(False)
    except Error:
        print("There was a problem connecting to the database, please make sure the database information is correct!")
    return con_try

def read_file(path, conn): #ENTRY POINT
    try:
        ending = path.split('.')[-1].lower()
    except:
        print("Error finding file ending")
        return
    if ending == "json":
        acc = parse_file_json(path)
    elif ending == "xlsx":
        acc = parse_file_xlsx(path)
    elif ending == "csv":
        acc = parse_file_csv(path)
    else:
        print ("Invalid file format")
        return
    write_accounts(acc, conn)


if __name__ == "__main__":
    read_file("dummy_data/accounts_shifted.xlsx", connect_mysql("dbkey.json"))
