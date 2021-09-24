import xml.etree.ElementTree as eTree
import json
import traceback
import csv
from typing import List
import jaydebeapi
from jaydebeapi import Error
from openpyxl import load_workbook
from openpyxl import worksheet
import datetime


class Transaction:
    def __init__(self):
            self.origin: int = None
            self.destination: int = None
            self.memo: str = None
            self.value: float = None
            self.time_stamp: datetime.datetime = None
            
class Card_Transaction:
    def __init__(self):
            self.card: int = None
            self.acc: int = None
            self.memo: str = None
            self.value: float = None
            self.pin: int = None
            self.cvc1: int = None
            self.cvc2: int = None
            self.location: str = None
            self.time_stamp: datetime.datetime = None

def check_null(string: str) -> str:
    if string == '\\N' or string == '':
        return None
    return string            

def parse_json_dict(json_dict: dict) -> Transaction:

    trans = Transaction()
    trans.origin = json_dict["origin_account"]
    trans.destination = json_dict["destination_account"]
    trans.memo = json_dict["memo"]
    trans.value = json_dict["transfer_value"]
    trans.time_stamp = json_dict["time_stamp"]
    return trans

def parse_json_dict_cards(json_dict: dict) -> Card_Transaction:

    trans = Card_Transaction()
    trans.card = json_dict["card_num"]
    trans.acc = json_dict["merchant_account_id"]
    trans.memo = json_dict["memo"]
    trans.value = json_dict["transfer_value"]
    trans.pin = check_null(json_dict["pin"])
    trans.cvc1 = check_null(json_dict["cvc1"])
    trans.cvc2 = check_null(json_dict["cvc2"])
    trans.location = json_dict["location"]
    trans.time_stamp = json_dict["time_stamp"]
    return trans

def parse_file_json(path) -> List:
    try:
        f = open(path, "r")
    except:
        print("error opening file")
        return None

    json_list = json.load(f)
    return_list = []
    for json_dict in json_list:
        try:
            return_list.append(parse_json_dict(json_dict))
        except:
            print("Line failed to be parsed")
    return return_list

def parse_file_json_cards(path) -> List:
    try:
        f = open(path, "r")
    except:
        print("error opening file")
        return None

    json_list = json.load(f)
    return_list = []
    for json_dict in json_list:
        try:
            return_list.append(parse_json_dict_cards(json_dict))
        except:
            print("Line failed to be parsed")
    return return_list

def parse_table_xlsx(ws: worksheet, bounds: tuple) -> List:
    ret_list = []
    for row in ws.iter_rows(min_row=bounds[0], min_col=bounds[1], max_col=bounds[1]+7):
        
        try: #test to see if we've reached the end of our data
            if row[0].value == None:
                return ret_list
        except:
            return ret_list
        try:
            trans = Transaction()
            trans.origin = row[0].value
            trans.destination = row[1].value
            trans.memo = row[2].value
            trans.value = row[3].value
            trans.time_stamp = row[4].value
            ret_list.append(trans)
        except:
            print('Row failed to be parsed')
    return ret_list

def parse_table_xlsx_cards(ws: worksheet, bounds: tuple) -> List:
    ret_list = []
    for row in ws.iter_rows(min_row=bounds[0], min_col=bounds[1], max_col=bounds[1]+8):
        
        try: #test to see if we've reached the end of our data
            if row[0].value == None:
                return ret_list
        except:
            return ret_list
        try:
            trans = Card_Transaction()
            trans.card = row[0].value
            trans.acc = row[1].value
            trans.memo = row[2].value
            trans.value = row[3].value
            trans.pin = check_null(row[4].value)
            trans.cvc1 = check_null(row[5].value)
            trans.cvc2 = check_null(row[6].value)
            trans.location = row[7].value
            trans.time_stamp = row[8].value
            ret_list.append(trans)
        except:
            print('Row failed to be parsed')
    return ret_list

def find_xlsx_bounds(ws: worksheet, num_columns: int): #num_columns should be the number of columns we care about i.e. not id
    row_num = 0
    for row in ws.iter_rows(min_row=1, max_row=20):
        row_num += 1 #yes this is an odd pattern, but it lets me use the iterator and report back the row number
        for i in range(0, 10):
            try:
                if row[i].value == "id": #header and primary key
                    return (row_num+1, i+2)# row_num is already 1 indexed, but we want to add one because we hit id. For column we add one for 0-index and one to get from id to accounts_id
                elif row[i].value == "origin_account" or row[i].value == "card_num": #header but no primary key
                    return (row_num+1, i+1)#adjust row_num for headers and i for 0-index
                elif row[i].value != '' and row[i].value != None: #if we hit data
                    try:
                        if row[i+num_columns].value == '' or row[i+num_columns].value == None: #if this col is empty we assume it doesn't have primary key
                            return(row_num, i+1) #adjust i for 0-index
                        else:
                            return(row_num, i+2) #adjust i for 0-index and primary key
                    except: #means there's probably not anything there
                            return(row_num, i+1) #adjust i for 0-index
            except:
                pass #we don't care about an out of range here
    return None

def parse_file_xlsx(path: str) -> List:
    wb = load_workbook(filename=path, read_only=True)
    ws = None
    for sheet in wb: #tries to find a sheet named accounts
        if sheet.title.lower() == "transactions":
            ws = sheet
    if ws == None:
        ws = wb.active #gets the first sheet
    bounds = find_xlsx_bounds(ws, 5)
    return parse_table_xlsx(ws, bounds)

def parse_file_xlsx_cards(path: str) -> List:
    wb = load_workbook(filename=path, read_only=True)
    ws = None
    for sheet in wb: #tries to find a sheet named accounts
        if sheet.title.lower() == "card_transactions":
            ws = sheet
    if ws == None:
        ws = wb.active #gets the first sheet
    bounds = find_xlsx_bounds(ws, 9)
    return parse_table_xlsx_cards(ws, bounds)

def date_to_string(date): #differs from str(date) in that it accepts none
    if date:
        return str(date)
    return None

def write_transactions(trans_list: List, conn: jaydebeapi.Connection) -> None:
    query = 'INSERT INTO transactions(origin_account, destination_account, memo, transfer_value, time_stamp) VALUES (?,?,?,?,?)'
    cur = conn.cursor()
    for trans in trans_list:
        vals = (trans.origin, trans.destination, trans.memo, trans.value, date_to_string(trans.time_stamp))
        try:
            cur.execute(query, vals)
        except:
            print("could not write transaction with values " + str(vals))

def write_transactions_cards(trans_list: List, conn: jaydebeapi.Connection) -> None:
    query = 'INSERT INTO card_transactions(card_num, merchant_account_id, memo, transfer_value, pin, cvc1, cvc2, location, time_stamp) VALUES (?,?,?,?,?,?,?,?,?)'
    cur = conn.cursor()
    for trans in trans_list:
        vals = (trans.card, trans.acc, trans.memo, trans.value, trans.pin, trans.cvc1, trans.cvc2, trans.location, date_to_string(trans.time_stamp))
        try:
            cur.execute(query, vals)
        except:
            print("could not write transaction with values " + str(vals))

def parse_file_csv(path: str) -> List:
    with open(path, newline="") as acc_file:
        reader = csv.reader(acc_file, delimiter=',')
        row_count = 0
        ret_list = [] #return list
        for row in reader:
            # Functionality for ingesting transactions
            try:
                trans = Transaction()
                trans.origin = row[0]
                trans.destination = row[1]
                trans.memo = row[2]
                trans.value = row[3]
                trans.time_stamp = row[4]
                ret_list.append(trans)
            except:
                print("Could not add transaction on line " + str(row_count) + ": " + str(row))
                continue
    return ret_list

def parse_file_csv_cards(path: str) -> List:
    with open(path, newline="") as acc_file:
        reader = csv.reader(acc_file, delimiter=',')
        row_count = 0
        ret_list = [] #return list
        for row in reader:
            # Functionality for ingesting transactions
            try:
                trans = Card_Transaction()
                trans.card = row[0]
                trans.acc = row[1]
                trans.memo = row[2]
                trans.value = row[3]
                trans.pin = check_null(row[4])
                trans.cvc1 = check_null(row[5])
                trans.cvc2 = check_null(row[6])
                trans.location = row[7]
                trans.time_stamp = row[8]
                ret_list.append(trans)
            except:
                print("Could not add transaction on line " + str(row_count) + ": " + str(row))
                continue
    return ret_list

def parse_file_xml(path: str) -> List:
    ret_list = []
    try:
        tree = eTree.parse(path)
        root = tree.getroot()
        for child in root:
            try:
                trans = Transaction()
                trans.origin = child.find('origin_account').text
                trans.destination = child.find('destination_account').text
                trans.memo = child.find('memo').text
                trans.value = child.find('transaction_value').text
                trans.time_stamp = child.find('time_stamp').text
                ret_list.append(trans)
            except:
                print("Could not add transaction")
                print("Skipping line...\n")
    except eTree.ParseError:
        traceback.print_exc()
    return ret_list

def parse_file_xml_cards(path: str) -> List:
    ret_list = []
    try:
        tree = eTree.parse(path)
        root = tree.getroot()
        for child in root:
            try:
                trans = Card_Transaction()
                trans.card = child.find('card_num').text
                trans.acc = child.find('merchant_account_id').text
                trans.memo = child.find('memo').text
                trans.value = child.find('transaction_value').text
                trans.pin = check_null(child.find('pin').text)
                trans.cvc1 = check_null(child.find('cvc1').text)
                trans.cvc2 = check_null(child.find('cvc2').text)
                trans.location = child.find('location').text
                trans.time_stamp = child.find('time_stamp').text
                ret_list.append(trans)
            except:
                print("Could not add transaction")
                print("Skipping line...\n")
    except eTree.ParseError:
        traceback.print_exc()
    return ret_list

def connect_mysql(path):
    con_try = None
    try:
        f = open(path, 'r')
        key = json.load(f)        
        con_try = jaydebeapi.connect(key["driver"], key["location"], key["login"], key["jar"] )
        con_try.jconn.setAutoCommit(False)
    except Error:
        print("There was a problem connecting to the database, please make sure the database information is correct!")
    return con_try

def read_file(path, conn): #ENTRY POINT
    try:
        ending = path.split('.')[-1].lower()
    except:
        print("Error finding file ending")
        return
    if ending == "json":
        trans = parse_file_json(path)
    elif ending == "xlsx":
        trans = parse_file_xlsx(path)
    elif ending == "csv":
        trans = parse_file_csv(path)
    else:
        print ("Invalid file format")
        return
    write_transactions(trans, conn)

def read_file_cards(path, conn): #ENTRY POINT
    try:
        ending = path.split('.')[-1].lower()
    except:
        print("Error finding file ending")
        return
    if ending == "json":
        trans = parse_file_json_cards(path)
    elif ending == "xlsx":
        trans = parse_file_xlsx_cards(path)
    elif ending == "csv":
        trans = parse_file_csv_cards(path)
    else:
        print ("Invalid file format")
        return
    write_transactions_cards(trans, conn)


if __name__ == "__main__":
    conn =  connect_mysql("dbkey.json")
    read_file_cards("dummy_data/card_transactions.xlsx", conn)
    conn.commit()
